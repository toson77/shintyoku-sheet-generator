from typing import NamedTuple
import openpyxl as excel
import calendar
import itertools
from openpyxl.styles import PatternFill 
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter


class GenSheet():
    def __init__(self, legends: NamedTuple, filename: str, month: int, year: int = 2021):
        # excel file gen
        self.__wb = excel.Workbook()
        self.__legends = legends

        file_name: str = filename
        self.__year: str = str(year)
        self.__month: str = str(month)

        # [[0,1,2,3],[4,5,6,7]] -> [0,1,2,3,4,5,6,7] ->[1,2,3,4,5,6,7]
        days = calendar.monthcalendar(
            int(self.__year), int(self.__month))
        days = list(itertools.chain.from_iterable(days))
        self.__days = list(filter(lambda x: x != 0, days))

        if ".xlsx" not in file_name:
            self.__wbname: str = file_name + ".xlsx"
        else:
            self.__wbname: str = file_name

    def __make_sheet(self, key: str, value: tuple) -> None:
        # get sheet
        ws = self.__wb[key]
        # set name
        ws.cell(row=1, column=1).value = self.__month + "月"
        for index, val in enumerate(value):
            ws.cell(row=1, column=index+3).value = val
            ws.cell(row=1, column=index+3).alignment = Alignment(horizontal='center', vertical='center')
        # set date
        for index, val in enumerate(self.__days):
            week_flg = self.__str_week(
                int(self.__year), int(self.__month), val)
            ws.cell(row=index+2, column=1).value = str(val) + "日"
            ws.cell(row=index+2, column=2).value = week_flg
        # fixed A1
        ws.freeze_panes = 'A2'

    def __str_week(self, year: int, month: int, day: int) -> str:
        flag = calendar.weekday(year, month, day)
        day_tuple: tuple[str, ...] = (
            '(月)', '(火)', '(水)', '(木)', '(金)', '(土)', '(日)')
        return day_tuple[flag]

    def __fill_background_color(self, key: str) -> None:
        # backgroundcolor
        ws = self.__wb[key]
        for index, row in enumerate(ws):
            for cell in row:
                fill = PatternFill(patternType='solid', fgColor='d3d3d3')
                if ws[cell.coordinate].value == '(土)' or ws[cell.coordinate].value == '(日)':
                    ws[cell.coordinate].fill = fill
                    ws[ws.cell(row=index+1, column=1).coordinate].fill = fill

    def __set_sizeof_cell(self, key: str) -> None:
        ws = self.__wb[key]
        # set high
        for index, row in enumerate(ws):
            ws.row_dimensions[index+2].height = 99.75
        # set width
        for row in ws:
            for index, cell in enumerate(row):
                if index < 2:
                    ws.column_dimensions[get_column_letter(index+1)].width = 8.38
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    ws.column_dimensions[get_column_letter(index+1)].width = 30



    def gen(self) -> None:
        # add sheet
        for legend in self.__legends._asdict():
            self.__wb.create_sheet(title=legend)
        # make sheet
        for legend in self.__legends._asdict().items():
            self.__make_sheet(legend[0], legend[1])
        # set bgcolor
        for legend in self.__legends._asdict():
            self.__fill_background_color(legend)
        # set wide and high of cell
        for legend in self.__legends._asdict():
            self.__set_sizeof_cell(legend)
        # delete init sheet
        self.__wb.remove(self.__wb['Sheet'])
        self.__wb.save(self.__wbname)
